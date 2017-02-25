import openpyxl
"data_only set to True for extracting the results Excel provides after formula"
wb2 = openpyxl.load_workbook('StoryTR_mx.xlsx',data_only=True)
TR = wb2['Sheet2']
"iter_cols return a generator"
run01index = [n.value for m in TR.iter_cols(min_row = 2,min_col = 2, max_col = 2, max_row = 17) for n in m ]
run01TRonset = [n.value for m in TR.iter_cols(min_row = 2,min_col = 3, max_col = 3, max_row = 17) for n in m ]
run01TRoffset = [n.value for m in TR.iter_cols(min_row = 2,min_col = 4, max_col = 4, max_row = 17) for n in m ]

run02index = [n.value for m in TR.iter_cols(min_row = 2,min_col = 7, max_col = 7, max_row = 17) for n in m ]
run02TRonset = [n.value for m in TR.iter_cols(min_row = 2,min_col = 8, max_col = 8, max_row = 17) for n in m ]
run02TRoffset = [n.value for m in TR.iter_cols(min_row = 2,min_col = 9, max_col = 9, max_row = 17) for n in m ]

run03index = [n.value for m in TR.iter_cols(min_row = 2,min_col = 12, max_col = 12, max_row = 17) for n in m ]
run03TRonset = [n.value for m in TR.iter_cols(min_row = 2,min_col = 13, max_col = 13, max_row = 17) for n in m ]
run03TRoffset = [n.value for m in TR.iter_cols(min_row = 2,min_col = 14, max_col = 14, max_row = 17) for n in m ]

run04index = [n.value for m in TR.iter_cols(min_row = 2,min_col = 17, max_col = 17, max_row = 17) for n in m ]
run04TRonset = [n.value for m in TR.iter_cols(min_row = 2,min_col = 18, max_col = 18, max_row = 17) for n in m ]
run04TRoffset = [n.value for m in TR.iter_cols(min_row = 2,min_col = 19, max_col = 19, max_row = 17) for n in m ]

run05index = [n.value for m in TR.iter_cols(min_row = 2,min_col = 22, max_col = 22, max_row = 17) for n in m ]
run05TRonset = [n.value for m in TR.iter_cols(min_row = 2,min_col = 23, max_col = 23, max_row = 17) for n in m ]
run05TRoffset = [n.value for m in TR.iter_cols(min_row = 2,min_col = 24, max_col = 24, max_row = 17) for n in m ]

stirun01 = []
for n in range(len(run01index)):
    stirun01[run01TRonset[n]:run01TRoffset[n]+1] = [run01index[n]] * len(range(run01TRonset[n],run01TRoffset[n]+1))
"remove the first 6 volumes"
stirun01 = stirun01[6:]
stirun02 = []
for n in range(len(run02index)):
    stirun02[run02TRonset[n]:run02TRoffset[n]+1] = [run02index[n]] * len(range(run02TRonset[n],run02TRoffset[n]+1))
stirun02 = stirun02[6:]

stirun03 = []
for n in range(len(run03index)):
    stirun03[run03TRonset[n]:run03TRoffset[n]+1] = [run03index[n]] * len(range(run03TRonset[n],run03TRoffset[n]+1))
stirun03 = stirun03[6:]
stirun04 = []
for n in range(len(run04index)):
    stirun04[run04TRonset[n]:run04TRoffset[n]+1] = [run04index[n]] * len(range(run04TRonset[n],run04TRoffset[n]+1))
stirun04 = stirun04[6:]
stirun05 = []
for n in range(len(run05index)):
    stirun05[run05TRonset[n]:run05TRoffset[n]+1] = [run05index[n]] * len(range(run05TRonset[n],run05TRoffset[n]+1))
stirun05 = stirun05[6:]
sti1D = stirun01 + stirun02 + stirun03 + stirun04 + stirun05

"seperate the 1D file into 4 stimuli 1D"
"Normal Story 1D"
NS1D = [1 if x == 1 else 0 for x in sti1D ]
" Connected Sentences 1D"
CS1D = [1 if x == 2 else 0 for x in sti1D ]
"Unconneceted Sentences 1D"
US1D = [1 if x == 3 else 0 for x in sti1D ]
"Scrambled Words 1D"
SW1D = [1 if x == 4 else 0 for x in sti1D ]


"write list in to file"
with open('NS1D.mx','w') as file:
    for item in NS1D:
        file.write("%d\n" % item)
with open('CS1D.mx','w') as file:
    for item in CS1D:
        file.write("%d\n" % item)
with open('US1D.mx','w') as file:
    for item in US1D:
        file.write("%d\n" % item)
with open('SW1D.mx','w') as file:
    for item in SW1D:
        file.write("%d\n" % item)
